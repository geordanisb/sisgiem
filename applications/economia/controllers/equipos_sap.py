# -*- coding: utf-8 -*-
# try something like
def removeNonAscii(s): return "".join(
    i for i in s if ord(i) < 128)  # para remover los caracteres basuras del archivo CSV en los campos q haga falta
def to_name(texto):
    texto = IS_SLUG()(texto)[0]
    texto = texto.replace("-", "_")
    return texto+"_"
db.equipo_sap.id.readable = False
@cache(request.env.path_info, time_expire=5, cache_model=cache.ram)
@auth.requires_membership("editor")
def index():
    grid = SQLFORM.grid(db.equipo_sap)
    btn_importar = A(I(_class="glyphicon glyphicon-import"),_href=URL("importar"),_class="btn btn-default",_title="Importar Équipos SAP")
    btn_migrar = A(I(_class="glyphicon glyphicon-transfer"), _href=URL("migrar"), _class="btn btn-default",_title="Migrar Équipos SAP")
    btn_exportar = A(I(_class="glyphicon glyphicon-save"), _href=URL("exportar"), _class="btn btn-default",
                   _title="Exportar Équipos SAP para carga inicial")

    grid.elements("div.web2py_grid ")[0].insert(0, btn_exportar)
    grid.elements("div.web2py_grid ")[0].insert(0, btn_migrar)
    grid.elements("div.web2py_grid ")[0].insert(0, btn_importar)
    return locals()

def migrar():
    count = 0
    query = db.activo.n_inventario == db.equipo_sap.n_inventario_
    query &= db.activo.mon == "USD"
    try:
        rows = db(query).select(db.equipo_sap.id,db.equipo_sap.equipo_,db.activo.n_inventario,db.activo.denominacin_del_activo_fijo,db.activo.actfijo,db.activo.local_,db.activo.fecapit,db.activo.id)
        for r in rows:
            uo = db(db.uo.local_ == r["activo"].local_).select().first()
            if uo:
                responsable = Responsable(uo.responsable,cargo=False)
                equipo_sap = db.equipo_sap(r["equipo_sap"].id)
                actfijo = r["activo"].actfijo
                local = uo.local_
                equipo_sap.update_record(campo_de_clasificacin_ = responsable,actfijo_ = actfijo, local_ = local)
                count += 1
    except Exception, e:
                redirect(URL('error','index', vars=dict(error=str(e))))
    else: redirect(URL('index'))
    return locals()

def importar():
    import csv, os
    
    path = os
    columns = None
    uploadfolder = os.path.join(request.folder, 'uploads')

    form = SQLFORM.factory(
        Field('file', 'upload', label="Archivo CSV",
              requires=IS_NOT_EMPTY(error_message="Seleccione el archivo (csv)."), uploadfolder=uploadfolder))

    if form.process().accepted:
        try:
            path = os.path.join(request.folder, 'uploads', 'equipos_sap.csv')
            path_temp = os.path.join(request.folder, 'uploads', 'equipos_sap_temp.csv')

            if os.path.isfile(path_temp):
                os.remove(path_temp)
            os.rename(os.path.join(uploadfolder, form.vars.file), os.path.join(uploadfolder, 'equipos_sap_temp.csv'))

            if os.path.isfile(path):
                os.remove(path)

            with open(path_temp,"rb") as f, open(path,"wb") as f1:
                reader = csv.reader(f,delimiter=",")
                cabecera = reader.next()

                while "Equipo" not in cabecera:
                    cabecera = reader.next()
                if "" in cabecera: raise Exception("La tabla tiene columnas vacías. ajuste las columnas según los datos.")


                columns = [to_name(removeNonAscii(r)) for r in cabecera]
                rows = []
                db.equipo_sap.truncate('RESTART IDENTITY CASCADE')
                for r in reader:
                    record = dict(zip(columns,r))
                    if record["equipo_"]:
                        db.equipo_sap.insert(**record)
        except Exception, e:
            redirect(URL('error', 'index', vars=dict(error=str(e))))
        else:
            redirect(URL('index'))
    return locals()

def exportar():
    import csv, os
    try:
        rows = db(db.equipo_sap.id>0).select()
        path = os.path.join(request.folder,"uploads","equipos_sap_exported.csv")

        with open(path, "wb") as f:

            writer = csv.writer(f, delimiter=",")

            for r in rows:
                # activo = db(db.activo.n_inventario==r.n_inventario_).select().first()
                # if activo:
                row = ["", r.denominacin_de_objeto_tcnico_, "Clase de objeto", "Grupo autorizaciones", r.n_inventario_,
                       r.ce_, r.emplaz_, r.are_, "", "Indicador ABC", "ET01", r.div_, r.cecoste_, "E031", "001", \
                       "TECINFO", "Perfil de catálogo", "NO DE SERIE", r.local_, r.campo_de_clasificacin_, r.actfijo_,
                       r.modelo_del_equipo_, r.marca_del_equipo_, r.selllo_]
                writer.writerow(row)

    except Exception, e:
        redirect(URL('error', 'index', vars=dict(error=str(e))))
    else:
        return response.stream(path, request=request,filename="equipos_sap_exported.csv", attachment=True)

def g():
    # Import `load_workbook` module from `openpyxl`
    from openpyxl import load_workbook

    # Load in the workbook
    wb = load_workbook('ver.xlsx')

    # Get sheet names
    #r = wb.get_sheet_names()

    sheet = wb.get_sheet_by_name('Hoja1')
    # Retrieve the value of a certain cell
    a2 = sheet['A2'].value
    b1 = sheet['B1']
    row = b1.row

    # Retrieve the column letter of your element
    column = b1.column

    # Retrieve the coordinates of the cell
    cordenadas = b1.coordinate

    return locals()